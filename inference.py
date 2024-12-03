import os
import torch
import numpy as np
from models.model import Model
import yaml
import pandas as pd
from torchvision import transforms as T
import skimage.io as io 
from util import remove_aggregations
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from scipy import ndimage
from Dataset import MAUI_transform
from sklearn.metrics import f1_score

# Load inference configuration from YAML file
with open("config/config_inference.yaml", "r") as f:
    cfg = yaml.safe_load(f)

# Define the output folder for the results based on the configuration
outputs_folder = 'results/{}'.format(cfg['results_folder_name'])

# Determine the device for computation (GPU if available, else CPU)
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

# Load data from the compression matrix form
protein_name = cfg["protein"]

# Define a transformation for center cropping (if specified in the config)
center_crop_transform = T.Compose([T.CenterCrop(cfg['center_crop'])]) if cfg['center_crop'] else T.Compose([])

# Create a list of paths to trained model checkpoints
checkpoint_path = 'pretrained_models/{}.pt'.format(cfg['model_name'])

# Initialize an empty list to store trained model instances
models_list = []

model = Model(model_type=cfg['model_type'], in_channels=1, n_classes=1, filters=cfg['model_features'], is_bias=cfg['bias']).to(device).eval()
checkpoint = torch.load(checkpoint_path, map_location=lambda storage, loc: storage)
model.load_state_dict(checkpoint["model"])
models_list.append(model)

# Create an empty DataFrame to store F1 scores
f1_df = pd.DataFrame(columns=(protein_name, f'{protein_name} (dilated)', 'clean_signal_sparsity', 'clean_median_val','FPR', 'FNR'), dtype=float)

# Load FOV data and process it
def load_images(fov_path_raw, fov_path_clean):
    raw_images, clean_images = [], []
    raw_image_filepath = os.path.join(fov_path_raw, f'{protein_name}.tif')
    clean_image_filepath = os.path.join(fov_path_clean, f'{protein_name}.tif')

    try:
        raw_images.append(io.imread(raw_image_filepath))
        clean_images.append(io.imread(clean_image_filepath))
    except FileNotFoundError:
        print(f"File not found: {raw_image_filepath} or {clean_image_filepath}")
        return [], []

    raw_images = torch.from_numpy(np.stack(raw_images).astype(np.float32)).unsqueeze(0).to(device)
    clean_images = torch.from_numpy(np.stack(clean_images).astype(np.float32)).unsqueeze(0).to(device)

    # center crop
    raw_images = center_crop_transform(raw_images)
    clean_images = center_crop_transform(clean_images)

    return raw_images, clean_images


def min_max(image):
    # Calculate the percentile
    q = 0.99999  # Assuming 'percentile' is a parameter in your configuration
 
    # Calculate the q-th percentile
    perc_value = torch.kthvalue(image.flatten(), int(image.numel() * q)).values
    
    # Apply the normalization equation
    normalized_tensor = image / (1.1 * perc_value)
    
    # Cap values greater than 1
    normalized_tensor[normalized_tensor > 1] = 1
    
    return normalized_tensor, perc_value

def anscombe(image):
    # Applying the Anscombe transform
    anscombe_image = 2 * torch.sqrt(image + 3/8)
    return anscombe_image

def prediction(models_list, raw_images):
    with torch.no_grad():
        if cfg['image_normalization'] == 'min_max':
            raw_images_normalized, perc_value = min_max(raw_images)
            model_output = (models_list[0](raw_images_normalized)['preds'] > 0.5).float()
        elif cfg['image_normalization'] == 'anscombe':
            raw_images_normalized = anscombe(raw_images)
            model_output = (models_list[0](raw_images_normalized)['preds'] > 0.5).float()
        else:
            if cfg['confidence_output']:
                model_output = (models_list[0](raw_images)['preds']).float()
            else:
                model_output = (models_list[0](raw_images)['preds'] > 0.5).float()
        
        # model_output = torch.from_numpy(remove_aggregations(model_output[0].cpu().numpy())).to(device).unsqueeze(0)
        return raw_images * model_output.detach()

# Save FOV images and results
def save_images(outputs_folder, fov, clean_images, preds, raw_images):
    # create fov dir
    fov_dir_path = os.path.join(os.path.join(outputs_folder, fov))
    os.makedirs(fov_dir_path, exist_ok=True)

    # save clean image (GT)
    io.imsave(os.path.join(fov_dir_path, '{}_clean.tif'.format(protein_name)), clean_images[0,0].cpu().numpy().astype('uint8'), check_contrast=False)
    # save predicted clean image
    io.imsave(os.path.join(fov_dir_path, '{}_pred.tif'.format(protein_name)), preds[0, 0].cpu().numpy().astype('uint8'), check_contrast=False)
    # save raw image
    io.imsave(os.path.join(fov_dir_path, '{}_raw.tif'.format(protein_name)), raw_images[0, 0].cpu().numpy().astype('uint8'), check_contrast=False)


# Analyze FOV results and calculate F1 scores
def analyze_result(outputs_folder, fov, clean_image, pred, f1_df):
    clean_image = clean_image[0,0].cpu()
    pred = pred[0,0].cpu()

    # Binarize the filtered images
    clean_image_binary = (clean_image.flatten() > 0)
    pred_image_binary = (pred.flatten() > 0)

    # calculate F1 score
    f1_score_result = f1_score(clean_image_binary, pred_image_binary, zero_division=1)
    f1_df.at[fov, protein_name] = f1_score_result

    # Calculate True Positives, False Positives, and False Negatives
    tp_mask = ((clean_image > 0) * (pred > 0)).float()
    fp_mask = ((clean_image == 0) * (pred > 0)).int()
    fn_mask = ((clean_image > 0) * (pred == 0)).int()

    # Calculate FPR and store in FPR_df
    fpr = fp_mask.sum() / (fp_mask.sum() + fn_mask.sum())
    f1_df.at[fov, 'FPR'] = fpr.item()

    # Calculate FNR and store in FNR_df
    fnr = fn_mask.sum() / (fp_mask.sum() + fn_mask.sum())
    f1_df.at[fov, 'FNR'] = fnr.item()

    # Calculate sparsity of clean signal
    sparsity_val = clean_image_binary.sum()
    f1_df.at[fov, 'clean_signal_sparsity'] = sparsity_val.item()

    # Calculate sparsity of clean signal
    if len(clean_image[clean_image > 0]):
        median_val = clean_image[clean_image > 0].median()
        f1_df.at[fov, 'clean_median_val'] = median_val.item()

    # Dilate prediction for dilated F1 score
    dilated_tp_mask = (ndimage.uniform_filter(tp_mask, size=5) > 0).astype(int)
    dilated_pred = (pred > 0).int() - fp_mask * dilated_tp_mask + fn_mask * dilated_tp_mask
    pred_dilated = dilated_pred.flatten()

    # Calculate dilated F1 score
    dilated_F1_score_result = f1_score(clean_image_binary, pred_dilated, zero_division=1)
    f1_df.at[fov, f'{protein_name} (dilated)'] = dilated_F1_score_result

    # rename reconstruction file with its F1 score
    os.rename(os.path.join(outputs_folder, fov, '{}_pred.tif'.format(protein_name)), 
                os.path.join(outputs_folder, fov, '{}_pred_{:.3f}_{:.3f}.tif'.format(protein_name, dilated_F1_score_result, f1_score_result)))


# Iterate over the test dataset and perform inference
for sub_dataset in os.listdir(os.path.join(cfg['test_dataset_path'], 'noisy')):
    noisy_sub_dataset_path = os.path.join(cfg['test_dataset_path'], 'noisy', sub_dataset)
    clean_sub_dataset_path = os.path.join(cfg['test_dataset_path'], 'clean', sub_dataset)
    if os.path.isdir(noisy_sub_dataset_path) and sub_dataset in cfg['sub_datasets']:
        print(f'Sub dataset {sub_dataset}:')
        for fov in os.listdir(noisy_sub_dataset_path):
            fov_path_raw = os.path.join(noisy_sub_dataset_path, fov, 'TIFs')
            fov_path_clean = os.path.join(clean_sub_dataset_path, fov, 'TIFs')
            if os.path.isdir(fov_path_raw):
                print(f'FOV {fov}:')
                raw_images, clean_images = load_images(fov_path_raw, fov_path_clean)
                if len(raw_images) == 0:
                    continue
                preds = prediction(models_list, raw_images)
                save_images(outputs_folder, fov, clean_images, preds, raw_images)
                analyze_result(outputs_folder, fov, clean_images, preds, f1_df)
                print(f'Inference process for FOV {fov} is done!\n')

f1_df = f1_df.sort_values(by=f1_df.columns[1], ascending=False).round(3)

# Save the calculated F1 scores to a CSV file
f1_df.to_csv(os.path.join(outputs_folder, 'F1 scores results - {}.csv').format(cfg['results_folder_name']))
output_file_path = os.path.join(outputs_folder, 'F1 scores results - {}.xlsx').format(cfg['results_folder_name'])
f1_df.to_excel(output_file_path)

# Load the workbook and select the active sheet
workbook = load_workbook(output_file_path)
sheet = workbook.active

# Iterate over all columns except the index column (starting from the second column)
for col in range(1, sheet.max_column + 1):
    max_length = max(len(str(cell.value)) for cell in sheet[get_column_letter(col)])
    sheet.column_dimensions[get_column_letter(col)].width = max_length + 2

# Apply a white-to-green color scale to each column
for col in range(2, sheet.max_column + 1):  # start from 2 to skip the index column
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='FFFFFF',  # White for the minimum
        end_type='max', end_color='009900'  # Green for the maximum
    )
    col_letter = chr(64 + col) if col <= 26 else chr(64 + (col - 1) // 26) + chr(64 + (col - 1) % 26 + 1)
    sheet.conditional_formatting.add(f'{col_letter}2:{col_letter}{sheet.max_row}', color_scale_rule)

# Enable the filter functionality for the columns
sheet.auto_filter.ref = sheet.dimensions

# Save the changes to the workbook
workbook.save(output_file_path)

#########
import seaborn as sns
import pandas as pd
import matplotlib.pyplot as plt
import os
import matplotlib as mpl
mpl.rcParams['pdf.fonttype'] = 42
mpl.rc('font', family='arial')
mpl.rc('font', serif='arial')

# set run name
run_name = cfg['results_folder_name']

F1_scores_path = f'results/{run_name}/F1 scores results - {run_name}.csv'

# load df of the F1 scores table
F1_scores_df = pd.read_csv(F1_scores_path, index_col=0, usecols=[0,1,2])
# Reorder the columns in the data frame by median
medians = F1_scores_df.median().sort_values(ascending=False)
F1_scores_df = F1_scores_df[medians.index]

# Create the box plot
num_channels = medians.shape[0]
plt.figure(figsize=(num_channels * 2 + 2, 30), dpi=300)
sns.set(style="white")
PROPS = {
'boxprops':{'facecolor':'grey', 'edgecolor':'black'},
'medianprops':{'color':'black'},
'whiskerprops':{'color':'black'},
'capprops':{'color':'black'}
}
ax = sns.boxplot(data=F1_scores_df, color='w', width=0.6, showfliers=True, linewidth=0.5, **PROPS)
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)
ax.spines['bottom'].set_linewidth(0.5)
ax.spines['left'].set_linewidth(0.5)

# Add median lines to the plot
medians = F1_scores_df.median().round(2)
for i, median in enumerate(medians):
    plt.text(i, median + 0.01, f"{median:.2f}", horizontalalignment='center', fontsize=24, color='black')

# Set the x-tick labels to channel names
plt.xticks(range(0, num_channels), F1_scores_df.columns, fontsize=30, rotation=90)
plt.yticks(fontsize=30)

# Add labels and title to the plot
plt.xlabel("\nProteins", fontsize=24)
plt.ylabel("F1 Scores\n", fontsize=24)
plt.ylim(0, 1)
plt.title("F1 Scores by protein\n", fontsize=30)

# Show the plot
plt.savefig(os.path.join(os.path.dirname(F1_scores_path), f'Box Plot - F1 scores - {run_name}.pdf'), format='pdf',bbox_inches='tight')
plt.show()